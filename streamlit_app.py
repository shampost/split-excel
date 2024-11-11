import csv
import math
import os
import tempfile
from io import BytesIO
from zipfile import ZipFile

import chardet
import pandas as pd
import streamlit as st
import xlrd
import xlwt
from openpyxl import Workbook, load_workbook

def count_rows(file_path, file_extension):
    """Count actual data rows (excluding header) for different file formats."""
    if file_extension in ['.csv', '.txt']:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            total_lines = sum(1 for _ in f)
        return total_lines - 1  # Exclude header
    elif file_extension == '.xlsx':
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        total_rows = ws.max_row - 1  # Exclude header
        wb.close()
        return total_rows
    elif file_extension == '.xls':
        wb = xlrd.open_workbook(file_path, on_demand=True)
        sheet = wb.sheet_by_index(0)
        total_rows = sheet.nrows - 1  # Exclude header
        wb.release_resources()
        del wb  # Ensure resources are freed
        return total_rows
    else:
        raise ValueError("Unsupported file format")

@st.cache_data(show_spinner=False)
def get_encoding_and_delimiter(file_path):
    """Detect the encoding and delimiter of a CSV or TXT file."""
    with open(file_path, 'rb') as f:
        sample = f.read(100000)
    result = chardet.detect(sample)
    detected_encoding = result['encoding'] if result['confidence'] > 0.8 else 'utf-8'

    with open(file_path, 'r', encoding=detected_encoding, errors='replace') as f:
        sample_lines = [f.readline() for _ in range(5)]
    sample_str = ''.join(sample_lines)

    try:
        delimiter = csv.Sniffer().sniff(sample_str).delimiter
    except csv.Error:
        delimiter = ','
    return detected_encoding, delimiter

def process_csv_txt(file_path, file_extension, row_count_split, num_files, data_rows, progress_bar):
    """Process CSV and TXT files in chunks."""
    detected_encoding, delimiter = get_encoding_and_delimiter(file_path)

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        reader = pd.read_csv(
            file_path,
            chunksize=row_count_split,
            encoding=detected_encoding,
            sep=delimiter,
            engine='python',
            dtype=str
        )

        chunks_processed = 0
        for i, chunk in enumerate(reader):
            buffer = BytesIO()
            chunk.to_csv(
                buffer,
                index=False,
                encoding=detected_encoding,
                sep=delimiter,
                lineterminator='\n'
            )
            buffer.seek(0)
            filename = f"split_file_{i+1}{file_extension}"
            zip_file.writestr(filename, buffer.read())
            buffer.close()
            chunks_processed += len(chunk)
            progress_bar.progress(min(chunks_processed / data_rows, 1.0))

    return zip_buffer

def process_xlsx(file_path, file_extension, row_count_split, num_files, data_rows, progress_bar):
    """Process XLSX files without loading entire file into memory."""
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        row_buffer = []
        file_index = 1
        total_rows_processed = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_buffer.append(row)
            total_rows_processed += 1

            if len(row_buffer) == row_count_split or total_rows_processed == data_rows:
                chunk_wb = Workbook()
                chunk_ws = chunk_wb.active
                chunk_ws.append(header)
                for data_row in row_buffer:
                    chunk_ws.append(data_row)

                buffer = BytesIO()
                chunk_wb.save(buffer)
                buffer.seek(0)
                filename = f"split_file_{file_index}{file_extension}"
                zip_file.writestr(filename, buffer.read())
                buffer.close()
                chunk_wb.close()
                row_buffer = []
                file_index += 1
                progress_bar.progress(min(total_rows_processed / data_rows, 1.0))

        wb.close()
    return zip_buffer

def process_xls(file_path, file_extension, row_count_split, num_files, data_rows, progress_bar):
    """Process XLS files without loading entire file into memory."""
    wb = xlrd.open_workbook(file_path, on_demand=True)
    sheet = wb.sheet_by_index(0)
    header = sheet.row_values(0)

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        file_index = 1
        total_rows_processed = 0

        for start_row in range(1, sheet.nrows, row_count_split):
            end_row = min(start_row + row_count_split, sheet.nrows)
            chunk_wb = xlwt.Workbook()
            chunk_ws = chunk_wb.add_sheet('Sheet1')

            # Write header
            for col_num, value in enumerate(header):
                chunk_ws.write(0, col_num, value)

            # Write data rows
            for row_num, row_idx in enumerate(range(start_row, end_row), start=1):
                row_values = sheet.row_values(row_idx)
                for col_num, value in enumerate(row_values):
                    chunk_ws.write(row_num, col_num, value)
                total_rows_processed += 1

            buffer = BytesIO()
            chunk_wb.save(buffer)
            buffer.seek(0)
            filename = f"split_file_{file_index}{file_extension}"
            zip_file.writestr(filename, buffer.read())
            buffer.close()
            del chunk_wb
            file_index += 1
            progress_bar.progress(min(total_rows_processed / data_rows, 1.0))

    wb.release_resources()
    del wb
    return zip_buffer

def main():
    st.title("File Splitter")

    # Initialize session state
    if 'processed' not in st.session_state:
        st.session_state.processed = False
    if 'zip_buffer' not in st.session_state:
        st.session_state.zip_buffer = None
    if 'last_row_count' not in st.session_state:
        st.session_state.last_row_count = None
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = None

    file = st.file_uploader(
        "Choose a CSV, XLSX, XLS, or TXT file.",
        type=['csv', 'xlsx', 'xls', 'txt']
    )

    if file is not None:
        try:
            # Reset state if new file
            if st.session_state.uploaded_file_name != file.name:
                st.session_state.processed = False
                st.session_state.zip_buffer = None
                st.session_state.uploaded_file_name = file.name

            row_count_split = st.number_input(
                "Enter the number of rows per split file",
                min_value=1,
                value=400000,
                key="row_count"
            )

            # Reset state if row count changes
            if st.session_state.last_row_count != row_count_split:
                st.session_state.processed = False
                st.session_state.zip_buffer = None

            st.session_state.last_row_count = row_count_split

            file_extension = os.path.splitext(file.name)[1].lower()

            # Save uploaded file to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_input_file:
                tmp_input_file.write(file.getbuffer())
                tmp_input_file_path = tmp_input_file.name

            # Get accurate row count excluding header
            data_rows = count_rows(tmp_input_file_path, file_extension)

            # Adjust the number of files calculation
            num_files = (data_rows + row_count_split - 1) // row_count_split  # Ceiling division

            st.write(f"Number of files to be created: {num_files}")

            if st.button("Confirm and Split File"):
                st.write("Splitting file... this might take some time.")
                progress_bar = st.progress(0)

                if file_extension in ['.csv', '.txt']:
                    zip_buffer = process_csv_txt(tmp_input_file_path, file_extension, row_count_split, num_files, data_rows, progress_bar)
                elif file_extension == '.xlsx':
                    zip_buffer = process_xlsx(tmp_input_file_path, file_extension, row_count_split, num_files, data_rows, progress_bar)
                elif file_extension == '.xls':
                    zip_buffer = process_xls(tmp_input_file_path, file_extension, row_count_split, num_files, data_rows, progress_bar)
                else:
                    st.error("Unsupported file format.")
                    return

                st.session_state.processed = True
                st.session_state.zip_buffer = zip_buffer
                zip_buffer.seek(0)

                # Remove temporary input file
                os.remove(tmp_input_file_path)

        except Exception as e:
            st.error(f"An error occurred during file splitting: {e}")

    # Display download button
    if st.session_state.processed and st.session_state.zip_buffer is not None:
        st.download_button(
            label="Download All Split Files",
            data=st.session_state.zip_buffer.getvalue(),
            file_name="split_files.zip",
            mime='application/zip'
        )

if __name__ == "__main__":
    main()